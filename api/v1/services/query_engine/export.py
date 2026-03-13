"""
Generacion de reportes CSV, Excel y PDF para consultas del Query Builder.
Optimizado para datasets grandes: CSV usa streaming, Excel usa write_only,
PDF pre-calcula anchos una sola vez.
"""
from io import BytesIO, StringIO
from datetime import datetime
from collections.abc import Generator
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF


# ── CSV (streaming por chunks) ───────────────────────────────────────

_CSV_CHUNK_SIZE = 1000


def generate_csv_streaming(rows: list[dict], columns_meta: list[dict]) -> Generator[bytes, None, None]:
    """Genera CSV en chunks para StreamingResponse. No carga todo en RAM."""
    headers = [c["label"] for c in columns_meta]
    keys = [c["column_name"] for c in columns_meta]

    # BOM + header
    buf = StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(headers)
    yield b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")

    # Datos en chunks
    for i in range(0, len(rows), _CSV_CHUNK_SIZE):
        chunk_buf = StringIO()
        chunk_writer = csv.writer(chunk_buf, lineterminator="\n")
        for row in rows[i:i + _CSV_CHUNK_SIZE]:
            chunk_writer.writerow([row.get(k, "") for k in keys])
        yield chunk_buf.getvalue().encode("utf-8")


# ── Excel (write-only mode para menor uso de RAM) ────────────────────

def generate_excel(rows: list[dict], columns_meta: list[dict], title: str = "Consulta") -> BytesIO:
    """Genera Excel (.xlsx) en modo write_only para eficiencia con datasets grandes."""
    headers = [c["label"] for c in columns_meta]
    keys = [c["column_name"] for c in columns_meta]

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=title[:31])

    # En write_only no se pueden aplicar estilos por celda directamente,
    # pero si se pueden pasar listas de celdas con estilo.
    from openpyxl.cell import WriteOnlyCell

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Header row con estilos
    header_cells = []
    for h in headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_cells.append(cell)
    ws.append(header_cells)

    # Data rows (sin estilos individuales para velocidad)
    for row in rows:
        ws.append([row.get(k, "") for k in keys])

    # Anchos de columna estimados (basado en header + muestreo)
    sample_size = min(100, len(rows))
    for col_idx, key in enumerate(keys, 1):
        max_len = len(headers[col_idx - 1])
        for row in rows[:sample_size]:
            val = row.get(key, "")
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF ──────────────────────────────────────────────────────────────

class _QueryPDF(FPDF):
    def __init__(self, title: str, col_widths: list[float], col_headers: list[str], *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._report_title = title
        self._col_widths = col_widths
        self._col_headers = col_headers

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, self._report_title, new_x="LMARGIN", new_y="NEXT", align="C")
        self.set_font("Helvetica", "", 9)
        self.cell(
            0, 6,
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            new_x="LMARGIN", new_y="NEXT", align="C",
        )
        self.ln(4)
        # Re-dibujar header de tabla en cada pagina
        self._draw_table_header()

    def _draw_table_header(self):
        self.set_font("Helvetica", "B", 7)
        self.set_fill_color(31, 78, 121)
        self.set_text_color(255, 255, 255)
        for i, h in enumerate(self._col_headers):
            label = h[:20] + "..." if len(h) > 20 else h
            self.cell(self._col_widths[i], 8, label, border=1, fill=True, align="C")
        self.ln()
        self.set_font("Helvetica", "", 7)
        self.set_text_color(0, 0, 0)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="C")


def generate_pdf(rows: list[dict], columns_meta: list[dict], title: str = "Consulta") -> BytesIO:
    """Genera PDF landscape. Header de tabla se repite en cada pagina."""
    headers = [c["label"] for c in columns_meta]
    keys = [c["column_name"] for c in columns_meta]

    # Max 10 columnas por espacio
    max_cols = min(len(headers), 10)
    headers = headers[:max_cols]
    keys = keys[:max_cols]

    # Anchos proporcionales
    usable_width = 277
    col_width = usable_width / len(headers)
    col_widths = [col_width] * len(headers)

    pdf = _QueryPDF(
        title, col_widths, headers,
        orientation="L", unit="mm", format="A4",
    )
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Datos
    for idx, row in enumerate(rows):
        if idx % 2 == 1:
            pdf.set_fill_color(235, 241, 247)
        else:
            pdf.set_fill_color(255, 255, 255)

        for i, key in enumerate(keys):
            val = str(row.get(key, ""))
            if len(val) > 25:
                val = val[:22] + "..."
            pdf.cell(col_widths[i], 7, val, border=1, fill=True, align="C")
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf
