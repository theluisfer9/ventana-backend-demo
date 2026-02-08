"""
Generacion de reportes de beneficiarios en Excel y PDF.
"""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fpdf import FPDF


# ── Columnas del reporte ─────────────────────────────────────────────

COLUMNS = [
    "ID Hogar",
    "CUI Jefe",
    "Nombre Jefe Hogar",
    "Sexo Jefe",
    "Departamento",
    "Municipio",
    "Lugar Poblado",
    "Area",
    "No. Personas",
    "IPM",
    "Clasif. IPM",
    "PMT",
    "Clasif. PMT",
]

# Columnas para PDF (sin PMT por espacio)
PDF_COLUMNS = COLUMNS[:11]


def _row(b: dict) -> list:
    """Convierte un beneficiario RSH a fila de reporte."""
    return [
        b.get("hogar_id", ""),
        b.get("cui_jefe_hogar", ""),
        b.get("nombre_completo", ""),
        "Femenino" if b.get("sexo_jefe_hogar") == "F" else "Masculino",
        b.get("departamento", ""),
        b.get("municipio", ""),
        b.get("lugar_poblado", ""),
        b.get("area", ""),
        b.get("numero_personas", 0),
        round(b.get("ipm_gt", 0), 4),
        b.get("ipm_gt_clasificacion", ""),
        round(b.get("pmt", 0), 4),
        b.get("pmt_clasificacion", ""),
    ]


# ── Excel ────────────────────────────────────────────────────────────

def generate_excel(rows: list[dict]) -> BytesIO:
    """Genera un archivo .xlsx en memoria con los beneficiarios dados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiarios"

    # Estilos de header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Escribir headers
    for col_idx, title in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir datos
    for row_idx, beneficiario in enumerate(rows, 2):
        values = _row(beneficiario)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-width
    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = len(str(COLUMNS[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Congelar primera fila
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF ──────────────────────────────────────────────────────────────

class _BeneficiarioPDF(FPDF):
    """PDF landscape con header/footer personalizados."""

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, "Reporte de Beneficiarios", new_x="LMARGIN", new_y="NEXT", align="C")
        self.set_font("Helvetica", "", 9)
        self.cell(
            0, 6,
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            new_x="LMARGIN", new_y="NEXT", align="C",
        )
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="C")


def generate_pdf(rows: list[dict]) -> BytesIO:
    """Genera un archivo PDF landscape en memoria."""
    pdf = _BeneficiarioPDF(orientation="L", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Anchos de columna para 11 columnas (landscape A4 ~277mm usable)
    col_widths = [20, 22, 50, 18, 34, 34, 30, 18, 14, 14, 28]  # total ~282 for 11 cols

    # Header de tabla
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    for i, col in enumerate(PDF_COLUMNS):
        pdf.cell(col_widths[i], 8, col, border=1, fill=True, align="C")
    pdf.ln()

    # Datos
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(0, 0, 0)

    for idx, beneficiario in enumerate(rows):
        values = _row(beneficiario)
        # Fila alternada
        if idx % 2 == 1:
            pdf.set_fill_color(235, 241, 247)
            fill = True
        else:
            pdf.set_fill_color(255, 255, 255)
            fill = True

        for i in range(len(PDF_COLUMNS)):
            val = str(values[i]) if values[i] is not None else ""
            # Truncar si es muy largo
            if len(val) > 35:
                val = val[:32] + "..."
            align = "C" if i in (0, 3, 7, 8, 9, 10) else "L"
            pdf.cell(col_widths[i], 7, val, border=1, fill=fill, align=align)
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf
